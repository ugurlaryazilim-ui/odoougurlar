from odoo import models, fields, api
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class CountSession(models.Model):
    _name = 'ugurlar.barcode.count.session'
    _description = 'Sayım Oturumu'
    _order = 'create_date desc'

    name = fields.Char('Sayım Kodu', default='Yeni', required=True, readonly=True, copy=False)
    location_id = fields.Many2one('stock.location', string='Raf (Konum)', required=True, readonly=True, states={'draft': [('readonly', False)]})
    user_id = fields.Many2one('res.users', string='Operatör', default=lambda self: self.env.uid, readonly=True, states={'draft': [('readonly', False)]})
    date = fields.Datetime('Tarih', default=fields.Datetime.now, readonly=True, states={'draft': [('readonly', False)]})
    
    state = fields.Selection([
        ('draft', 'Taslak'),
        ('done', 'Tamamlandı'),
        ('validated', 'Onaylandı')
    ], string='Durum', default='done', tracking=True)

    operation_ids = fields.One2many('ugurlar.barcode.operation', 'count_session_id', string='Sayım Detayları', readonly=True, states={'draft': [('readonly', False)], 'done': [('readonly', False)]})
    
    notes = fields.Text('Notlar')
    
    total_products = fields.Integer('Farklı Ürün Sayısı', compute='_compute_totals')
    total_counted_qty = fields.Float('Toplam Sayılan Adet', compute='_compute_totals')
    
    @api.depends('operation_ids', 'operation_ids.quantity')
    def _compute_totals(self):
        for rec in self:
            rec.total_products = len(rec.operation_ids.mapped('product_id'))
            rec.total_counted_qty = sum(rec.operation_ids.mapped('quantity'))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'Yeni') == 'Yeni':
                vals['name'] = self.env['ir.sequence'].next_by_code('ugurlar.barcode.count.session') or 'Yeni'
        return super().create(vals_list)

    def action_validate(self):
        for rec in self:
            rec.state = 'validated'
            
    def action_draft(self):
        for rec in self:
            rec.state = 'draft'

    # ═══════════════════════════════════════════════════════
    # EXCEL EXPORT
    # ═══════════════════════════════════════════════════════

    def _get_attribute_value(self, tmpl, attr_name, cache=None):
        """Template'tan belirli bir attribute değerini al (cache destekli)."""
        cache_key = (tmpl.id, attr_name)
        if cache is not None and cache_key in cache:
            return cache[cache_key]
        value = ''
        line = tmpl.attribute_line_ids.filtered(
            lambda l: l.attribute_id.name == attr_name
            and l.attribute_id.create_variant == 'no_variant'
        )
        if line:
            value = ', '.join(line[0].value_ids.mapped('name'))
        if cache is not None:
            cache[cache_key] = value
        return value

    def action_export_xlsx(self):
        """Seçili sayım fişlerini detaylı Excel olarak indir."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            from odoo.exceptions import UserError
            raise UserError('openpyxl kütüphanesi yüklü değil.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sayım Raporu'

        # Başlık stilleri
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Sütun başlıkları
        headers = [
            'Sayım Kodu', 'Tarih', 'Raf (Konum)', 'Operatör', 'Durum',
            'Ürün Kodu', 'Barkod', 'Ürün Adı', 'Marka', 'Sezon',
            'Varolan Adet', 'Sayılan Adet', 'Fark', 'Ürün Sayısı',
        ]
        col_widths = [16, 18, 30, 15, 14, 18, 18, 45, 18, 14, 14, 14, 10, 14]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Durum çevirisi
        state_map = {'draft': 'Taslak', 'done': 'Tamamlandı', 'validated': 'Onaylandı'}

        # Veri satırları
        row_idx = 2
        attr_cache = {}
        data_font = Font(name='Calibri', size=10)
        num_align = Alignment(horizontal='right')

        # Negatif fark için kırmızı, pozitif için yeşil
        red_font = Font(name='Calibri', size=10, color='CC0000', bold=True)
        green_font = Font(name='Calibri', size=10, color='006600', bold=True)

        for session in self.sorted(key=lambda s: s.date or s.create_date, reverse=True):
            # Sadece sayım operasyonlarını al (count_adjustment hariç)
            counting_ops = session.operation_ids.filtered(
                lambda o: o.operation_type == 'counting')

            if not counting_ops:
                continue

            for op in counting_ops:
                product = op.product_id
                tmpl = product.product_tmpl_id if product else None

                # Ürün kodu (default_code / internal reference)
                product_code = product.default_code or '' if product else ''
                barcode = op.barcode or ''
                product_name = product.display_name if product else op.barcode or ''
                marka = self._get_attribute_value(tmpl, 'Marka', attr_cache) if tmpl else ''
                sezon = self._get_attribute_value(tmpl, 'Sezon/Yıl', attr_cache) if tmpl else ''

                row_data = [
                    session.name,
                    (session.date or session.create_date).strftime('%Y-%m-%d %H:%M') if (session.date or session.create_date) else '',
                    session.location_id.complete_name or '',
                    session.user_id.name or '',
                    state_map.get(session.state, session.state or ''),
                    product_code,
                    barcode,
                    product_name,
                    marka,
                    sezon,
                    op.theoretical_qty,
                    op.quantity,
                    op.difference_qty,
                    session.total_products,
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    # Sayısal sütunlar sağa hizala
                    if col_idx >= 11:
                        cell.alignment = num_align

                # Fark sütununu renklendir
                fark_cell = ws.cell(row=row_idx, column=13)
                if op.difference_qty < 0:
                    fark_cell.font = red_font
                elif op.difference_qty > 0:
                    fark_cell.font = green_font

                row_idx += 1

        # Üst satırı sabitle
        ws.freeze_panes = 'A2'

        # Excel dosyasını oluştur
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_data = base64.b64encode(output.read())
        output.close()

        # Attachment oluştur ve indir
        filename = f'Sayim_Raporu_{fields.Date.today()}.xlsx'
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
