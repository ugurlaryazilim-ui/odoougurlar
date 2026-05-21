import base64
import io
import logging

from odoo import models, fields, api

_logger = logging.getLogger(__name__)


class StockWarehouseNebim(models.Model):
    """Nebim depo kodu eşleştirmesi için stock.warehouse uzantısı."""
    _inherit = 'stock.warehouse'

    nebim_warehouse_code = fields.Char(
        string='Nebim Depo Kodu',
        help='Nebim tarafındaki depo kodu (ör: "MERKEZ", "DEPO1")',
    )


class StockLocationQty(models.Model):
    """Stok konumlarında toplam ürün adedini gösteren alan."""
    _inherit = 'stock.location'

    total_quant_qty = fields.Float(
        string='Adet',
        compute='_compute_total_quant_qty',
        digits=(16, 0),
        help='Bu konum ve tüm alt konumlarındaki toplam ürün stok adedi',
    )

    def _compute_total_quant_qty(self):
        for loc in self:
            # parent_path ile tüm alt konumları dahil et (ör: "1/5/12/" → LIKE "1/5/12/%")
            if loc.parent_path:
                self.env.cr.execute("""
                    SELECT COALESCE(SUM(sq.quantity), 0)
                    FROM stock_quant sq
                    JOIN stock_location sl ON sl.id = sq.location_id
                    WHERE sl.parent_path LIKE %s
                """, [loc.parent_path + '%'])
                loc.total_quant_qty = self.env.cr.fetchone()[0]
            else:
                loc.total_quant_qty = sum(loc.quant_ids.mapped('quantity'))

    # ═══ RAF DETAY DOSYASI EXCEL EXPORT ═══
    def action_export_shelf_detail(self):
        """Seçili konumları veya tümünü Excel olarak dışa aktar (HamurLabs formatı)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise models.UserError("openpyxl kütüphanesi yüklü değil. 'pip install openpyxl' çalıştırın.")

        # Seçili kayıtlar veya tüm dahili konumlar
        if self:
            locations = self
        else:
            locations = self.search([('usage', '=', 'internal')], order='complete_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raf Detay Listesi"

        # ─── BAŞLIK STİLLERİ ───
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        # ─── SÜTUN BAŞLIKLARI (HamurLabs formatı) ───
        headers = [
            'Path', 'Name', 'Id', 'Unique Code', 'Global Slot',
            'Total Quantity', 'Type', 'Code', 'Is Pick',
            'Warehouse', 'Max Sku Count', 'Is Pool', 'Is Salable',
            'Is Shelf', 'Slot', 'Shelf Restriction', 'Pallet Restriction',
            'Is Countable', 'Is Reservable', 'Extra Shelf Life',
            'Width', 'Height', 'Length', 'Product Group',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ─── VERİ SATIRLARI ───
        usage_map = {
            'supplier': 'Supplier',
            'view': 'View',
            'internal': 'Normal',
            'customer': 'Customer',
            'inventory': 'Inventory',
            'transit': 'Transit',
            'production': 'Production',
        }

        row_fill_even = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
        data_align = Alignment(vertical='center')

        for row_idx, loc in enumerate(locations, 2):
            # Depo bilgisi
            warehouse_name = ''
            if loc.warehouse_id:
                warehouse_name = loc.warehouse_id.name or ''

            # Alt konum var mı? (Is Shelf)
            has_children = bool(loc.child_ids)

            # Slot — pozisyon bilgisi
            slot_parts = []
            if loc.posx: slot_parts.append(str(loc.posx))
            if loc.posy: slot_parts.append(str(loc.posy))
            if loc.posz: slot_parts.append(str(loc.posz))
            slot = '.'.join(slot_parts) if slot_parts else ''

            # Toplam stok adedi (alt konumlar dahil)
            total_qty = loc.total_quant_qty

            # Unique product count in location
            product_count = len(loc.quant_ids.filtered(lambda q: q.quantity > 0).mapped('product_id'))

            row_data = [
                loc.complete_name or '',       # Path
                loc.name or '',                 # Name
                loc.id,                         # Id
                loc.barcode or '',              # Unique Code
                slot or 'ALL',                  # Global Slot
                total_qty,                      # Total Quantity
                usage_map.get(loc.usage, loc.usage or ''),  # Type
                loc.barcode or '',              # Code
                'Yes' if loc.usage == 'internal' and not has_children else 'No',  # Is Pick
                warehouse_name,                 # Warehouse
                product_count,                  # Max Sku Count (unique ürün sayısı)
                'No',                           # Is Pool
                'Yes' if loc.usage == 'internal' else 'No',  # Is Salable
                'Yes' if not has_children and loc.usage == 'internal' else 'No',  # Is Shelf
                slot or 'ALL',                  # Slot
                'ALL',                          # Shelf Restriction
                'ALL',                          # Pallet Restriction
                'Yes' if loc.usage == 'internal' else 'No',  # Is Countable
                'Yes' if loc.usage == 'internal' and not loc.scrap_location else 'No',  # Is Reservable
                0,                              # Extra Shelf Life
                0,                              # Width
                0,                              # Height
                0,                              # Length
                '',                             # Product Group
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = row_fill_even

        # ─── SÜTUN GENİŞLİKLERİ ───
        col_widths = {
            1: 40, 2: 20, 3: 8, 4: 16, 5: 12,
            6: 14, 7: 10, 8: 16, 9: 8,
            10: 20, 11: 14, 12: 8, 13: 10,
            14: 8, 15: 10, 16: 16, 17: 16,
            18: 12, 19: 12, 20: 14,
            21: 8, 22: 8, 23: 8, 24: 14,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # Autofilter
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

        # Freeze pane (başlık satırı)
        ws.freeze_panes = 'A2'

        # ─── DOSYA OLUŞTUR VE İNDİR ───
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())

        from datetime import datetime as dt
        filename = f"raf_detay_listesi_{dt.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'store_fname': filename,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
