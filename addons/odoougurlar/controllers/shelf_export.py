"""Raf Detay Dosyası Excel Export Controller."""
import io
import logging
from datetime import datetime as dt

from odoo import http
from odoo.http import request, content_disposition

_logger = logging.getLogger(__name__)


class ShelfExportController(http.Controller):

    @http.route('/odoougurlar/shelf_detail_export', type='http', auth='user')
    def shelf_detail_export(self, **kw):
        """Tüm dahili stok konumlarını Excel olarak indir."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            _logger.error("openpyxl kütüphanesi bulunamadı!")
            return request.make_response(
                "openpyxl kütüphanesi bulunamadı",
                headers=[('Content-Type', 'text/plain')],
            )

        try:
            return self._generate_shelf_excel(openpyxl, Font, PatternFill,
                                              Alignment, Border, Side)
        except Exception as e:
            _logger.exception("Raf export hatası: %s", str(e))
            return request.make_response(
                f"Raf export hatası: {e}",
                headers=[('Content-Type', 'text/plain')],
            )

    def _generate_shelf_excel(self, openpyxl, Font, PatternFill,
                              Alignment, Border, Side):
        """Excel dosyası oluştur ve döndür."""
        env = request.env
        cr = env.cr

        # ─── 1. Warehouse map: lot_stock_id → warehouse_name ───
        cr.execute("""
            SELECT sw.lot_stock_id, sw.name
            FROM stock_warehouse sw
        """)
        wh_lot_map = dict(cr.fetchall())

        # ─── 2. Tüm dahili lokasyonları çek ───
        cr.execute("""
            SELECT id, complete_name, name, barcode, usage,
                   posx, posy, posz, scrap_location, parent_path,
                   location_id
            FROM stock_location
            WHERE usage = 'internal'
            ORDER BY complete_name
        """)
        locations = cr.fetchall()
        _logger.info("Raf export: %d lokasyon bulundu", len(locations))

        # ─── 3. Child count map ───
        cr.execute("""
            SELECT location_id, COUNT(*)
            FROM stock_location
            WHERE location_id IS NOT NULL
            GROUP BY location_id
        """)
        child_count_map = dict(cr.fetchall())

        # ─── 4. Quant toplam: her lokasyon için (parent_path ile hiyerarşi) ───
        cr.execute("""
            SELECT sq.location_id, SUM(sq.quantity)
            FROM stock_quant sq
            JOIN stock_location sl ON sq.location_id = sl.id
            WHERE sl.usage = 'internal' AND sq.quantity > 0
            GROUP BY sq.location_id
        """)
        quant_direct_map = dict(cr.fetchall())

        # ─── 5. Product count: her lokasyondaki unique ürün ───
        cr.execute("""
            SELECT sq.location_id, COUNT(DISTINCT sq.product_id)
            FROM stock_quant sq
            WHERE sq.quantity > 0
            GROUP BY sq.location_id
        """)
        product_count_map = dict(cr.fetchall())

        # ─── Hiyerarşik toplam hesapla (parent_path) ───
        loc_data = {}
        for row in locations:
            loc_id = row[0]
            parent_path = row[9] or ''
            loc_data[loc_id] = {
                'row': row,
                'parent_path': parent_path,
            }

        def calc_total_qty(loc_id):
            """Bir lokasyon ve tüm alt lokasyonlarının toplam miktarı."""
            pp = loc_data[loc_id]['parent_path']
            total = 0.0
            for other_id, other in loc_data.items():
                if other['parent_path'].startswith(pp):
                    total += quant_direct_map.get(other_id, 0)
            return total

        # ─── Lokasyon → Warehouse eşleme ───
        def find_warehouse(parent_path):
            """parent_path üzerinden warehouse bul."""
            parts = [int(p) for p in parent_path.strip('/').split('/') if p]
            for lot_id, wh_name in wh_lot_map.items():
                if lot_id in parts:
                    return wh_name
            return ''

        # ─── EXCEL OLUŞTUR ───
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raf Detay Listesi"

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E4057', end_color='2E4057',
                                  fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        row_fill_even = PatternFill(start_color='F5F7FA', end_color='F5F7FA',
                                    fill_type='solid')
        data_align = Alignment(vertical='center')

        headers = [
            'Path', 'Name', 'Id', 'Unique Code', 'Global Slot',
            'Total Quantity', 'Type', 'Code', 'Is Pick',
            'Warehouse', 'Max Sku Count', 'Is Pool', 'Is Salable',
            'Is Shelf', 'Slot', 'Shelf Restriction', 'Pallet Restriction',
            'Is Countable', 'Is Reservable', 'Extra Shelf Life',
            'Width', 'Height', 'Length', 'Product Group',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        usage_map = {
            'supplier': 'Supplier', 'view': 'View', 'internal': 'Normal',
            'customer': 'Customer', 'inventory': 'Inventory',
            'transit': 'Transit', 'production': 'Production',
        }

        for row_idx, row in enumerate(locations, 2):
            (loc_id, complete_name, name, barcode, usage,
             posx, posy, posz, scrap, parent_path, parent_id) = row

            warehouse = find_warehouse(parent_path or '')
            has_children = child_count_map.get(loc_id, 0) > 0
            total_qty = calc_total_qty(loc_id)
            prod_count = product_count_map.get(loc_id, 0)

            slot_parts = []
            if posx: slot_parts.append(str(posx))
            if posy: slot_parts.append(str(posy))
            if posz: slot_parts.append(str(posz))
            slot = '.'.join(slot_parts) if slot_parts else ''

            row_data = [
                complete_name or '',
                name or '',
                loc_id,
                barcode or '',
                slot or 'ALL',
                total_qty,
                usage_map.get(usage, usage or ''),
                barcode or '',
                'Yes' if usage == 'internal' and not has_children else 'No',
                warehouse,
                prod_count,
                'No',
                'Yes' if usage == 'internal' else 'No',
                'Yes' if not has_children and usage == 'internal' else 'No',
                slot or 'ALL',
                'ALL',
                'ALL',
                'Yes' if usage == 'internal' else 'No',
                'Yes' if usage == 'internal' and not scrap else 'No',
                0, 0, 0, 0, '',
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = row_fill_even

        # Sütun genişlikleri
        col_widths = {
            1: 40, 2: 20, 3: 8, 4: 16, 5: 12,
            6: 14, 7: 10, 8: 16, 9: 8,
            10: 20, 11: 14, 12: 8, 13: 10,
            14: 8, 15: 10, 16: 16, 17: 16,
            18: 12, 19: 12, 20: 14,
            21: 8, 22: 8, 23: 8, 24: 14,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
        ws.freeze_panes = 'A2'

        # Dosya döndür
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.read()

        filename = f"raf_detay_listesi_{dt.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

        headers_resp = [
            ('Content-Type',
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
            ('Content-Length', len(file_data)),
        ]
        return request.make_response(file_data, headers=headers_resp)
