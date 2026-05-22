"""Ürün Listesi Excel Export Controller."""
import io
import json
import logging
from datetime import datetime as dt

from odoo import http
from odoo.http import request, content_disposition

_logger = logging.getLogger(__name__)


class ProductListExportController(http.Controller):

    @http.route('/odoougurlar/product_list_export', type='http', auth='user')
    def product_list_export(self, domain=None, **kw):
        """Raflardaki ürünleri Excel olarak indir. domain ile filtreleme."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            _logger.error("openpyxl kütüphanesi bulunamadı!")
            return request.make_response(
                "openpyxl kütüphanesi bulunamadı",
                headers=[('Content-Type', 'text/plain')],
            )

        try:
            return self._generate_product_excel(
                openpyxl, Font, PatternFill, Alignment, Border, Side,
                domain_str=domain,
            )
        except Exception as e:
            _logger.exception("Ürün export hatası: %s", str(e))
            return request.make_response(
                f"Ürün export hatası: {e}",
                headers=[('Content-Type', 'text/plain')],
            )

    def _generate_product_excel(self, openpyxl, Font, PatternFill,
                                Alignment, Border, Side, domain_str=None):
        """Ürün listesi Excel dosyası oluştur."""
        env = request.env
        cr = env.cr

        # ─── 0. Domain parse: filtreli lokasyon ID'leri ───
        loc_filter_sql = ""
        loc_filter_params = ()
        if domain_str:
            try:
                domain = json.loads(domain_str)
                if domain:
                    loc_records = env['stock.location'].sudo().search(domain)
                    if loc_records:
                        loc_filter_sql = " AND sq.location_id IN %s"
                        loc_filter_params = (tuple(loc_records.ids),)
                        _logger.info("Ürün export: filtre uygulandı, %d lokasyon",
                                     len(loc_records))
            except Exception as e:
                _logger.warning("Domain parse hatası: %s, tümü indirilecek", e)

        # ─── 1. Quant + Product + Location verileri ───
        cr.execute(f"""
            SELECT
                pp.id AS product_id,
                pp.default_code,
                COALESCE(pt.name->>'tr_TR', pt.name->>'en_US',
                         pt.name->>'tr', '') AS product_name,
                pt.type,
                pp.barcode,
                pp.nebim_variant_code,
                pp.nebim_barcode,
                pt.nebim_code,
                pt.nebim_color_code,
                sq.quantity,
                sq.in_date,
                sq.write_date,
                sl.barcode AS loc_barcode,
                sl.complete_name AS loc_path,
                sl.name AS loc_name,
                pc.complete_name AS categ_name,
                pt.id AS template_id
            FROM stock_quant sq
            JOIN product_product pp ON sq.product_id = pp.id
            JOIN product_template pt ON pp.product_tmpl_id = pt.id
            JOIN stock_location sl ON sq.location_id = sl.id
            LEFT JOIN product_category pc ON pt.categ_id = pc.id
            WHERE sl.usage = 'internal'
              AND sq.quantity > 0
              {loc_filter_sql}
            ORDER BY sl.complete_name, pp.default_code
        """, loc_filter_params)
        quant_rows = cr.fetchall()
        _logger.info("Ürün export: %d ürün-raf kaydı bulundu", len(quant_rows))

        # ─── 2. Template attribute'leri (Marka, Cinsiyet, Sezon vb.) ───
        cr.execute("""
            SELECT
                ptal.product_tmpl_id,
                COALESCE(pa.name->>'tr_TR', pa.name->>'en_US', '') AS attr_name,
                STRING_AGG(
                    COALESCE(pav.name->>'tr_TR', pav.name->>'en_US', ''),
                    ', '
                ) AS attr_value
            FROM product_template_attribute_line ptal
            JOIN product_attribute pa ON ptal.attribute_id = pa.id
            JOIN product_template_attribute_value ptav
                ON ptav.attribute_line_id = ptal.id
            JOIN product_attribute_value pav ON ptav.product_attribute_value_id = pav.id
            WHERE pa.create_variant = 'no_variant'
            GROUP BY ptal.product_tmpl_id, pa.name, pa.id
        """)
        tmpl_attrs = {}
        for tmpl_id, attr_name, value in cr.fetchall():
            if tmpl_id not in tmpl_attrs:
                tmpl_attrs[tmpl_id] = {}
            tmpl_attrs[tmpl_id][attr_name or ''] = value

        def get_attr(tmpl_id, *names):
            attrs = tmpl_attrs.get(tmpl_id, {})
            for n in names:
                if n in attrs:
                    return attrs[n]
            return ''

        # ─── 3. Varyant-bazlı attribute'ler (Renk, Beden) ───
        # product_variant_combination: her varyantın kendi Renk ve Beden değeri
        cr.execute("""
            SELECT
                pvc.product_product_id,
                COALESCE(pa.name->>'tr_TR', pa.name->>'en_US', '') AS attr_name,
                COALESCE(pav.name->>'tr_TR', pav.name->>'en_US', '') AS attr_value
            FROM product_variant_combination pvc
            JOIN product_template_attribute_value ptav
                ON pvc.product_template_attribute_value_id = ptav.id
            JOIN product_attribute_value pav
                ON ptav.product_attribute_value_id = pav.id
            JOIN product_attribute pa ON pav.attribute_id = pa.id
        """)
        variant_attrs = {}  # product_id → {attr_name: attr_value}
        for prod_id, attr_name, attr_value in cr.fetchall():
            if prod_id not in variant_attrs:
                variant_attrs[prod_id] = {}
            variant_attrs[prod_id][attr_name or ''] = attr_value

        def get_variant_attr(product_id, *names):
            """Varyantın kendi attribute değerini bul (Renk, Beden)."""
            attrs = variant_attrs.get(product_id, {})
            for n in names:
                if n in attrs:
                    return attrs[n]
            return ''

        # ─── EXCEL OLUŞTUR ───
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ürün Listesi"

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1B4332', end_color='1B4332',
                                  fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        row_fill_even = PatternFill(start_color='F0F5F0', end_color='F0F5F0',
                                    fill_type='solid')
        data_align = Alignment(vertical='center')

        headers = [
            'Ana Ürün Kodu', 'Ürün Kodu', 'Adı', 'Ürün Tipi', 'Marka',
            'Barkod', 'Beden', 'Raf Miktarı', 'Sezon', 'Renk',
            'Depoya Son Geliş Tarihi', 'Raf Tekil Kodu', 'Raf Yolu',
            'Ürün ID', 'Son Raflama Zamanı', 'Raf Tipi', 'Raf Statüsü',
            'SKU', 'EAN', 'Barkodlar', 'Kategoriler',
            'Ürün Sezonu', 'Ürün Tipi', 'Ürün Division',
            'Ürün Kategorisi', 'Ürün Alt Departman', 'Ürün Sınıfı',
            'Ürün Cinsiyeti', 'Ürün Grubu', 'Maksimum Raf Miktarı',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        type_map = {
            'consu': 'Tüketim',
            'service': 'Hizmet',
            'product': 'Ürün',
            'combo': 'Combo',
        }

        for row_idx, row in enumerate(quant_rows, 2):
            (product_id, default_code, product_name, ptype,
             barcode, variant_code, nebim_barcode, nebim_code,
             color_code, quantity, in_date, write_date,
             loc_barcode, loc_path, loc_name, categ_name,
             template_id) = row

            # Template seviye attribute'ler
            marka = get_attr(template_id, 'Marka')
            sezon = get_attr(template_id, 'Sezon/Yıl')
            cinsiyet = get_attr(template_id, 'Cinsiyet')
            urun_grubu = get_attr(template_id, 'Ürün Grubu')

            # Varyant-bazlı Renk ve Beden (barkoda özel)
            renk = get_variant_attr(product_id, 'Renk')
            beden = get_variant_attr(product_id, 'Beden')
            # Fallback: nebim_variant_code'dan
            if not beden and variant_code:
                parts = variant_code.split('-')
                beden = parts[-1] if len(parts) >= 2 else variant_code

            row_data = [
                nebim_code or '',                        # Ana Ürün Kodu
                default_code or '',                      # Ürün Kodu (varyant)
                product_name or '',                      # Adı
                type_map.get(ptype, ptype or ''),         # Ürün Tipi
                marka,                                   # Marka
                barcode or '',                           # Barkod
                beden,                                   # Beden
                quantity or 0,                           # Raf Miktarı
                sezon,                                   # Sezon
                renk,                                    # Renk
                in_date.strftime('%Y-%m-%d %H:%M') if in_date else '',
                loc_barcode or '',                       # Raf Tekil Kodu
                loc_path or '',                          # Raf Yolu
                product_id,                              # Ürün ID
                write_date.strftime('%Y-%m-%d %H:%M') if write_date else '',
                'Normal',                                # Raf Tipi
                'free',                                  # Raf Statüsü
                default_code or '',                      # SKU
                barcode or '',                           # EAN
                nebim_barcode or '',                     # Barkodlar
                categ_name or '',                        # Kategoriler
                sezon,                                   # Ürün Sezonu
                type_map.get(ptype, ptype or ''),         # Ürün Tipi
                '',                                      # Ürün Division
                categ_name or '',                        # Ürün Kategorisi
                '',                                      # Ürün Alt Departman
                '',                                      # Ürün Sınıfı
                cinsiyet,                                # Ürün Cinsiyeti
                urun_grubu,                              # Ürün Grubu
                0,                                       # Maksimum Raf Miktarı
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = row_fill_even

        # Sütun genişlikleri
        col_widths = {
            1: 16, 2: 18, 3: 30, 4: 12, 5: 16,
            6: 18, 7: 10, 8: 12, 9: 12, 10: 14,
            11: 20, 12: 16, 13: 40,
            14: 10, 15: 20, 16: 10, 17: 10,
            18: 14, 19: 18, 20: 24, 21: 30,
            22: 12, 23: 12, 24: 14,
            25: 30, 26: 16, 27: 14,
            28: 12, 29: 16, 30: 16,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
        ws.freeze_panes = 'A2'

        # Dosya döndür
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.read()

        filename = f"urun_listesi_{dt.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

        headers_resp = [
            ('Content-Type',
             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
            ('Content-Length', len(file_data)),
        ]
        return request.make_response(file_data, headers=headers_resp)
