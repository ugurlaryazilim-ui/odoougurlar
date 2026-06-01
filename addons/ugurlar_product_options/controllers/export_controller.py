import io
import json
import logging

from odoo import http
from odoo.http import request, content_disposition

_logger = logging.getLogger(__name__)

try:
    import xlsxwriter
except ImportError:
    _logger.warning("xlsxwriter not found, using openpyxl fallback")
    xlsxwriter = None


class ProductExportController(http.Controller):
    """
    Ürün Varyantları — Temiz XLSX Export
    Metni Kaydır (wrap_text) KAPALI, düzgün hücre formatı.
    """

    @http.route('/ugurlar/product/export', type='http', auth='user', methods=['POST'], csrf=True)
    def export_products(self, **kwargs):
        data = json.loads(kwargs.get('data', '{}'))
        ids = data.get('ids', [])
        field_defs = data.get('fields', [])

        if not ids:
            return request.make_response(
                'No IDs provided',
                headers=[('Content-Type', 'text/plain')],
                status=400,
            )

        # Ürünleri oku
        field_names = [f['name'] for f in field_defs]
        products = request.env['product.product'].sudo().browse(ids).read(field_names)

        # XLSX oluştur
        output = io.BytesIO()

        if xlsxwriter:
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            sheet = workbook.add_worksheet('Ürünler')

            # Stiller — wrap_text KAPALI
            header_fmt = workbook.add_format({
                'bold': True,
                'bg_color': '#714B67',
                'font_color': '#FFFFFF',
                'border': 1,
                'font_size': 11,
                'text_wrap': False,
                'valign': 'vcenter',
            })
            text_fmt = workbook.add_format({
                'text_wrap': False,
                'font_size': 10,
                'valign': 'vcenter',
            })
            money_fmt = workbook.add_format({
                'text_wrap': False,
                'font_size': 10,
                'num_format': '#,##0.00',
                'valign': 'vcenter',
            })
            float_fmt = workbook.add_format({
                'text_wrap': False,
                'font_size': 10,
                'num_format': '#,##0.00',
                'valign': 'vcenter',
            })

            # Sütun genişlikleri
            col_widths = {
                'default_code': 22,
                'barcode': 18,
                'display_name': 45,
                'name': 45,
                'list_price': 14,
                'standard_price': 14,
                'qty_available': 12,
                'virtual_available': 12,
            }

            # Header
            for col_idx, fdef in enumerate(field_defs):
                sheet.write(0, col_idx, fdef.get('label', fdef['name']), header_fmt)
                width = col_widths.get(fdef['name'], 15)
                sheet.set_column(col_idx, col_idx, width)

            # Satır yüksekliği — sabit
            sheet.set_default_row(20)

            # Veriler
            for row_idx, product in enumerate(products, start=1):
                for col_idx, fdef in enumerate(field_defs):
                    fname = fdef['name']
                    ftype = fdef.get('type', 'char')
                    value = product.get(fname, '')

                    # Many2many / Many2one → display string
                    if isinstance(value, (list, tuple)):
                        if len(value) >= 2:
                            value = value[1]  # (id, name) tuple
                        elif len(value) == 0:
                            value = ''
                        else:
                            value = str(value[0])
                    elif value is False:
                        value = ''

                    if ftype in ('monetary', 'float'):
                        try:
                            sheet.write_number(row_idx, col_idx, float(value or 0), money_fmt if ftype == 'monetary' else float_fmt)
                        except (ValueError, TypeError):
                            sheet.write(row_idx, col_idx, str(value), text_fmt)
                    else:
                        sheet.write(row_idx, col_idx, str(value) if value else '', text_fmt)

            # Freeze header
            sheet.freeze_panes(1, 0)
            # AutoFilter
            if field_defs:
                sheet.autofilter(0, 0, len(products), len(field_defs) - 1)

            workbook.close()
        else:
            # openpyxl fallback
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = 'Ürünler'

            header_fill = PatternFill(start_color='714B67', end_color='714B67', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            no_wrap = Alignment(wrap_text=False, vertical='center')

            for col_idx, fdef in enumerate(field_defs, start=1):
                cell = ws.cell(row=1, column=col_idx, value=fdef.get('label', fdef['name']))
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = no_wrap

            for row_idx, product in enumerate(products, start=2):
                for col_idx, fdef in enumerate(field_defs, start=1):
                    fname = fdef['name']
                    value = product.get(fname, '')
                    if isinstance(value, (list, tuple)):
                        value = value[1] if len(value) >= 2 else ''
                    elif value is False:
                        value = ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = no_wrap

            ws.freeze_panes = 'A2'
            wb.save(output)

        output.seek(0)

        filename = f'urunler_filtre_{len(products)}.xlsx'
        headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
            ('Content-Length', len(output.getvalue())),
        ]

        return request.make_response(output.getvalue(), headers=headers)
